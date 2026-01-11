from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime
from pathlib import Path
import os
import shutil
from db.database import get_db
from schemas.common import CommonResponse
from .schema import ReservationCreate, ReservationUpdate, ReservationResponse
from . import crud
from models.models import Permission, User
from core.security import get_current_user
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from .util import get_color_by_int, year_to_yearcode
router = APIRouter()


@router.get("/", response_model=CommonResponse[List[ReservationResponse]])
async def get_reservations(
    skip: int = 0,
    limit: int = 100,
    filter: Optional[str] = None,
    user_id: Optional[int] = None,
    complex_id: Optional[int] = None,
    vendor_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    reservations = crud.get_reservations(
        db,
        skip=skip,
        limit=limit,
        user_id=user_id,
        filter=filter,
        complex_id=complex_id,
        vendor_id=vendor_id
    )
    return CommonResponse(data=[ReservationResponse.model_validate(r) for r in reservations])


@router.get("/by-month/{year}/{month}", response_model=CommonResponse[List[ReservationResponse]])
async def get_reservations_by_month(year: int, month: int, filter: Optional[str] = None, db: Session = Depends(get_db)):
    reservations = crud.get_reservations_by_month(db, year=year, month=month, filter=filter)
    return CommonResponse(data=[ReservationResponse.model_validate(r) for r in reservations])

@router.get("/{reservation_id}", response_model=CommonResponse[ReservationResponse])
async def get_reservation(reservation_id: int, db: Session = Depends(get_db)):
    reservation = crud.get_reservation(db, reservation_id=reservation_id)
    if not reservation:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reservation not found")
    return CommonResponse(data=ReservationResponse.model_validate(reservation))


@router.get("/cotis/{cotis}", response_model=CommonResponse[ReservationResponse])
async def get_reservation_by_cotis(cotis: str, db: Session = Depends(get_db)):
    reservation = crud.get_reservation_by_cotis(db, cotis=cotis)
    if not reservation:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reservation not found")
    return CommonResponse(data=ReservationResponse.model_validate(reservation))


@router.get("/search/{query}", response_model=CommonResponse[List[ReservationResponse]])
async def search_reservations(query: str, db: Session = Depends(get_db)):
    reservations = crud.search_reservations(db, query=query)
    return CommonResponse(data=[ReservationResponse.model_validate(r) for r in reservations])


@router.post("/", response_model=CommonResponse[ReservationResponse], status_code=status.HTTP_201_CREATED)
async def create_reservation(
    cotis: str = Form(...),
    reserved_at: date = Form(...),
    is_transfered: bool = Form(...),
    description: Optional[str] = Form(None),
    location: int = Form(...),
    vendor: int = Form(...),
    template: Optional[int] = Form(None),
    reservation_photo: Optional[UploadFile] = File(None), 
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)):
    # COTIS 중복 체크
    existing = crud.get_reservation_by_cotis(db, cotis=cotis)
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="COTIS 번호가 이미 존재합니다.")
    
    # 관련 엔티티 존재 확인
    from domains.complex import crud as complex_crud
    complex = complex_crud.get_complex(db, complex_id=location)
    if not complex:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="단지를 찾을 수 없습니다.")
    
    from domains.vendor import crud as vendor_crud
    vendor_obj = vendor_crud.get_vendor(db, vendor_id=vendor)
    if not vendor_obj:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="업체를 찾을 수 없습니다.")
    
    from domains.user import crud as user_crud
    user = user_crud.get_user(db, user_id=current_user.id)
    
    template_id = template
    if template:
        from domains.template import crud as template_crud
        template_obj = template_crud.get_template(db, template_id=template)
        if not template_obj:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="양식을 찾을 수 없습니다.")

    db_reservation = crud.create_reservation(db=db, reservation=ReservationCreate(cotis=cotis, reserved_at=reserved_at, is_transfered=is_transfered, description=description, location=location, vendor=vendor, template=template_id), user_id=current_user.id)

    # PDF 파일 저장
    if reservation_photo and reservation_photo.filename:
        # 업로드 디렉토리 생성
        upload_dir = Path("data/reservation_photos")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # 파일 저장
        file_path = upload_dir / f"rp_{db_reservation.id}.pdf"
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(reservation_photo.file, buffer)
    return CommonResponse(data=ReservationResponse.model_validate(db_reservation))


@router.put("/{reservation_id}", response_model=CommonResponse[ReservationResponse])
async def update_reservation(
    reservation_id: int,
    reservation: ReservationUpdate,
    db: Session = Depends(get_db)
):
    update_data = reservation.model_dump(exclude_unset=True)
    
    # COTIS 중복 체크 (다른 예약과 중복되지 않는지)
    if 'cotis' in update_data:
        existing = crud.get_reservation_by_cotis(db, cotis=update_data['cotis'])
        if existing and existing.id != reservation_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="COTIS already exists")
    
    # 관련 엔티티 존재 확인
    if 'location' in update_data and update_data['location'] is not None:
        from domains.complex import crud as complex_crud
        complex = complex_crud.get_complex(db, complex_id=update_data['location'])
        if not complex:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Complex not found")
    
    if 'vendor' in update_data and update_data['vendor'] is not None:
        from domains.vendor import crud as vendor_crud
        vendor_obj = vendor_crud.get_vendor(db, vendor_id=update_data['vendor'])
        if not vendor_obj:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vendor not found")
    
    if 'author' in update_data and update_data['author'] is not None:
        from domains.user import crud as user_crud
        user = user_crud.get_user(db, user_id=update_data['author'])
        if not user:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    
    if 'template' in update_data and update_data['template'] is not None:
        from domains.template import crud as template_crud
        template_obj = template_crud.get_template(db, template_id=update_data['template'])
        if not template_obj:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found")
    
    db_reservation = crud.update_reservation(
        db=db,
        reservation_id=reservation_id,
        **update_data
    )
    if not db_reservation:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reservation not found")
    return CommonResponse(data=ReservationResponse.model_validate(db_reservation))


@router.delete("/{reservation_id}", response_model=CommonResponse[dict])
async def delete_reservation(reservation_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.permission != Permission.ADMIN:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="관리자만 삭제할 수 있습니다.")
    success = crud.delete_reservation(db=db, reservation_id=reservation_id)
    if not success:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reservation not found")
    return CommonResponse(data={"deleted": True})


@router.get("/{reservation_id}/reservation-photo")
async def get_reservation_photo(reservation_id: int, db: Session = Depends(get_db)):
    reservation = crud.get_reservation(db, reservation_id=reservation_id)
    if not reservation:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reservation not found")
    file_path = Path(f"data/reservation_photos/rp_{reservation_id}.pdf")
    if not file_path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reservation photo not found")
    description = reservation.description[:50] if reservation.description else ""
    return FileResponse(path=str(file_path), filename=f"{year_to_yearcode(reservation.reserved_at.year)}{reservation.reserved_at.strftime('%m%d')}-접사 -{reservation.vendor.name}-{reservation.location.name}-{description}.pdf", media_type="application/pdf")

@router.get("/{reservation_id}/generate-certificate-template")
async def generate_certificate_template(reservation_id: int, db: Session = Depends(get_db)):
    reservation = crud.get_reservation(db, reservation_id=reservation_id)
    if not reservation:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reservation not found")
    if not reservation.template_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found")
    from domains.template import util as template_util
    from domains.template import crud as template_crud
    template = template_crud.get_template(db, template_id=reservation.template_id)
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found")
    template_util.generate_certificate_template(template, reservation)
    
    # 생성된 파일 경로
    file_path = Path(f"data/certificates_template/{reservation_id}.xlsx")
    if not file_path.exists():
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="File generation failed")
    
    # 파일 다운로드 응답
    return FileResponse(
        path=str(file_path),
        filename=f"{year_to_yearcode(reservation.reserved_at.year)}{reservation.reserved_at.strftime('%m%d')}-완료확인양식-{reservation.vendor.name}-{reservation.location.name}-{reservation.description[:50]}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.post("/{reservation_id}/complete", response_model=CommonResponse[ReservationResponse])
async def complete_reservation(
    reservation_id: int,
    completed_at: Optional[str] = Form(None),
    certificate: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    reservation = crud.get_reservation(db, reservation_id=reservation_id)
    if not reservation:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reservation not found")
    
    # 완료일 처리
    update_data = {}
    if completed_at:
        try:
            update_data['completed_at'] = datetime.strptime(completed_at, '%Y-%m-%d').date()
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid date format. Use YYYY-MM-DD")
    
    # PDF 파일 저장
    if certificate and certificate.filename:
        # 업로드 디렉토리 생성
        upload_dir = Path("data/certificates")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # 파일 저장
        file_path = upload_dir / f"{reservation_id}.pdf"
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(certificate.file, buffer)
    
    # 예약 정보 업데이트
    if update_data:
        db_reservation = crud.update_reservation(
            db=db,
            reservation_id=reservation_id,
            **update_data
        )
        if not db_reservation:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reservation not found")
        return CommonResponse(data=ReservationResponse.model_validate(db_reservation))
    
    return CommonResponse(data=ReservationResponse.model_validate(reservation))

@router.get("/{reservation_id}/generate-certificate")
async def generate_certificate(reservation_id: int, db: Session = Depends(get_db)):
    reservation = crud.get_reservation(db, reservation_id=reservation_id)
    if not reservation:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reservation not found")
    file_path = Path(f"data/certificates/{reservation_id}.pdf")
    if not file_path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Certificate not found")
    description = reservation.description[:50] if reservation.description else ""
    return FileResponse(path=str(file_path), filename=f"{year_to_yearcode(reservation.reserved_at.year)}{reservation.reserved_at.strftime('%m%d')}-완료확인-{reservation.vendor.name}-{reservation.location.name}-{description}.pdf", media_type="application/pdf")

@router.get("/get-reservations/{year}")
async def get_reservations_by_year(year: int, db: Session = Depends(get_db)) -> FileResponse:
    reservations = crud.get_reservations_by_year(db, year=year)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "예약 목록"
    sheet.cell(row=1, column=1, value="접수일")
    sheet.cell(row=1, column=2, value="접수번호") # COTIS
    sheet.cell(row=1, column=3, value="완료일")
    sheet.cell(row=1, column=4, value="업체")
    sheet.cell(row=1, column=5, value="현장명")
    sheet.cell(row=1, column=6, value="연락처") # 현장 연락처
    sheet.cell(row=1, column=7, value="접수/작업내용")
    
    for i, reservation in enumerate(reservations, start=2):
        if not reservation.is_transfered:
            fill_color = PatternFill(start_color=get_color_by_int(i), end_color=get_color_by_int(i), fill_type="solid")
        else:
            fill_color = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
        sheet.cell(row=i, column=1, value=reservation.reserved_at.strftime('%Y-%m-%d')).fill = fill_color
        sheet.cell(row=i, column=2, value=reservation.cotis).fill = fill_color
        sheet.cell(row=i, column=3, value=reservation.completed_at.strftime('%m.%d') if reservation.completed_at else "").fill = fill_color
        sheet.cell(row=i, column=4, value=reservation.vendor.name).fill = fill_color
        sheet.cell(row=i, column=5, value=reservation.location.name)
        sheet.cell(row=i, column=6, value=reservation.location.tel)
        sheet.cell(row=i, column=7, value=reservation.description)
    
    # 디렉토리가 없으면 생성
    output_dir = Path("data/reservations")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    file_path = output_dir / f"r{year}.xlsx"
    workbook.save(str(file_path))
    return FileResponse(path=str(file_path), filename=f"{year}년 LH보수지시내역.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")